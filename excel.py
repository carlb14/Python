from openpyxl import load_workbook, Workbook
import dumb_menu
from prompt_toolkit import prompt
from openpyxl.styles import Font, Alignment
from prompt_toolkit.completion import WordCompleter
from prompt_toolkit.history import InMemoryHistory
import datetime
import os.path

print('\nWelcome to Python Automated Tool\n')

# Function to get the worksheet and workbook from the user
def Requirements():
    while True:
        Path = input('\nEnter the Location or PATH of the file (e.g: /home/carl/workloads.xlsx): ')
        if not os.path.isfile(Path):
            print('\n\033[91mThe file is not found.\033[00m')
        else:
            wb = load_workbook(Path)
            break

    return wb, Path

# Function to provide autocompletion for user input
def autocomplete_input(prompt_text, options):
    completer = WordCompleter(options, ignore_case=True)
    history = InMemoryHistory()

    print(f'\n{prompt_text}\n')

    while True:
        user_input = prompt("Enter your search query or type 'none' to input manually and contact the Developer: ", completer=completer, history=history)
        if user_input.lower() == 'none':
            selected_option = input('Enter the text: ')
            return selected_option
        else:
            suggestions = [option for option in options if user_input.lower() in option.lower()]
            if suggestions:
                menu_entry_index = dumb_menu.get_menu_choice(suggestions)
                selected_option = suggestions[menu_entry_index]
                print('>', selected_option)
                return selected_option
            else:
                print(f"\033[91mNot found. Re-check your input.\033[00m")

# Function to get the company name with autocompletion
def CompanyNameD():
    options = [
        "AFTERSIX DESSET&COFFEE INC", "AMICI FOODSERVICE VENTURES INC", "ANTHEM SHOPPES INC", "ASHLAR SECURITY SERVICES", "BIR", "BISTRO AMERICANO CORP", "BREADTALK PHILIPPINES INC", "BU PARTNERS INC", "CAFE VIA MARE", "CAFFEFRANCE CORP.", "CARA MIA SALES AND SERVICES INC", "CDC RESOURCES, INC", "CFAL OASIS DEVELOPMENT CORPORATION", "CK OF ASIA INC", "COFFEE MASTERS INC", "DECATHLON PHILIPPINES INC", "ENCHANTED KONGDOM INC", "ESPRESSO CONCEPTS INC", "EST. NAMNAM, INC", "EXA FUEL MARKETING OPC", "FOOGY MOUNTAIN COOKHOUSE", "GENMAICHA TEA CORPORATION", "GINGERSNAPS MARKETING IL CONIGLIO BIANCO CORPORATION", "GOLDEN DONUTS INC", "GOLDEN DONUTS, INC", "GOLDEN TEA LEAF INC", "GT & T GASOLINE STATION", "GTI", "HAPPY LEMON GROUP PHILIPPINES INC", "HEALTHY OPTION CORPORATION", "HOLDINGS INC.", "ILAOLLAO", "INTERBIKE MARKETING CORPORATION", "INTERNATIONAL TOYWORLD, INC", "IPPUDO PHILIPPINES, INC", "IT'S TIME FOR TIMS COFFEE INC.", "JBPX FOODS INC.", "JOLLIBEE BAGUIO BONIFACIO NIDEL COUNTRY LINE FOODS CORPORATION", "JOLLIBEE FOODS CORP.", "JOLLIBEE FOODS CORPORATION", "JOLLY888 FOOD CORP.", "JWB GASULINA CORP.", "KAREILA MGNT CORP", "LE PATISSIER INC",
        "LOLOMBOY WATERWORKS ASSOCIATION, INC", "LUGGAGE TRADING SERVICES INC", "MAJOR SHOPPING MANAGEMENT CORPORATION", "MARIE GALANTE FOOD SERVICES INC.", "MARY GRACE FOODS INC", "MAX'S GROUP INC", "MCDELIGHT SOUTH INC.", "MERCURY DRUG", "MERALCO", "MULTISPORTS INC", "MUNICIPALITY OF BOCAUE", "NATIONAL BOOK STORE, INC", "NOODLERAMA GROUP INC", "NOODLERAMA GROUP INC.", "OMAKASE INC.", "ORTIGAS HOME DEPOT", "OUTLOOK RESIDENCES INC", "PEPER BROERS, INC", "PRETIOLAS PHILIPPINES, INC", "REBEL ALLIANCE FOOD SUPPLY, INC", "REVSTAR TRADING", "RICHPRIME SALES INCORPORATED", "ROTI BOY, INC", "RUSTAN COFFEE CORP.",         "RUSTAN COFFEE CORPORATION", "SAN ROQUE SHELL SERVICE STATION", "SEATTLE'S BEST COFFEE MASTER INC", "SPECIALTY FOOD RETAILERS INC", "SPECIALTY LIFESTYLE CONCEPTS, INC", "SUNDAYS EST CAFE AND RESTAURANT INC", "SUNDAYS EST CAFE AND RESTAURANT INC.", "SWITCHSTART INC", "TABLE BLACK KITCHEN CORP.", "TEAMBAKEMARKETING CORP", "THE FRENCH BAKER INC", "THE TWELVE 28TH INC", "TOBISTO FOODSO PHI INC", "TOBISTRO FOODS, INC", "UNIGLIPPLOBE TRAVELWARE CO., INC", "UNIMART", "URBAN TASTE INC", "VIVA INTERNATIONAL FOOD & RESTAURANT, INC", "WILD FLOUR BAKERY+CAFE CORP", "WILDFLOUR BAKERY+CAFE CORP"]

    return autocomplete_input("Company name:", options)

# Function to get the address with autocompletion
def AddressD():
    options = [
        "QUEZON CITY",
        "MAKATI CITY",
        "BAGUIO CITY",
        "GUIGUINTO BULACAN",
        "SANJUAN CITY",
        "PASIG CITY",
        "APALIT PAMPANGA",
        "MARIKINA CITY",
        "MARILAO BULACAN",
        "MARILAO EXTENTION OFFICE",
        "PLARIDEL BULACAN",
        "AYALA TRINOMA",
        "TAGUIG CITY",
        "MANDALUYONG CITY",
        "BOCAUE BULACAN",
        "SAN PEDRO LAGUNA",
        "ANTIPOLO RIZAL",
        "STA MARIA BULACAN",
        "LOLOMBOY BOCAUE BULACAN",
        "SANTAROSA LAGUNA",
        "ANTIPOLO CITY"
    ]

    return autocomplete_input("Address:", options)

# Function to get the description with autocompletion
def Description():
    options = ['MEDICINE', 'GASOLINE', 'MEALS', 'DRINKS', 'SNACKS', 'SUPPLIES',
               'SECURITY FEE', 'WATER', 'ELECTRICITY', '2550- Q', '2307', 'OFFICE SUPPLIES']

    return autocomplete_input("Description:", options)

# Function to convert input date string to a formatted date string
def DateTime(input_str):
    input_str = input_str.replace('.', '')
    month_number = int(input_str[:2])
    day = int(input_str[2:4])
    year = int(input_str[4:])
    if year >= 0 and year <= 21:
        year += 2000
    else:
        year += 1900
    year %= 100  # Get only the last two digits
    month_full_name = datetime.date(year, month_number, day).strftime("%B")
    month_abbr = datetime.date(year, month_number, day).strftime("%b")

    datetime_number = f"{day}-{month_abbr}-{year}"
    return datetime_number, input_str, month_number

def ChoiceSheet(month_number, wb):
    month_names = {
        '01': 'January',
        '02': 'February',
        '03': 'March',
        '04': 'April',
        '05': 'May',
        '06': 'June',
        '07': 'July',
        '08': 'August',
        '09': 'September',
        '10': 'October',
        '11': 'November',
        '12': 'December'
    }
    sheet_name = month_names.get(month_number)
    if sheet_name:
        ws = wb[sheet_name]
        return ws
    else:
        print('Not Found')

# Function to get the TIN (Taxpayer Identification Number)
def TIN_D():
    while True:
        tin_number = input("\nEnter TIN number: ")
        if tin_number == "":
            print("\n\033[91mFill the blank!\033[00m")
        else:
            TIN = tin_number
            return TIN
def Again():
    while True:   
        results = input('\nDo you want to create more (y/n)? ')
        if results.lower() == 'y':
            Input_details(input("\nEnter the date (052424 = May,24,2024): "))
            break
        else:
            print('\nGood Bye!')
            exit() 
            
# Function to input details and calculate Cost of Products & Services
def Input_details(input_str):
    datetime_number, input_str, month_number = DateTime(input_str)
    company_name = CompanyNameD()  
    address_name = AddressD()
    TIN = TIN_D()
    description_name = Description()
    TotalAmountPaid = float(input('\nEnter the Total Amount Paid: '))
    InputVat = float(input('\nEnter the Input Vat: '))
    CPS = TotalAmountPaid - InputVat
    ws = ChoiceSheet(str(month_number).zfill(2), wb)
    
    
    #COnfirm Details
    print(f'Date: {datetime_number}')
    print(f'Company: {company_name}')
    print(f'Address: {address_name}')
    print(f'TIN: {TIN}')
    print(f'Description: {description_name}')
    print(f'Total Amount Paid: {TotalAmountPaid}')
    print(f'Input Vat: {InputVat}')
    print(f'Cost of Product & Services: {CPS}')
    
    
    
    while True:
        y_or_n = input('\nDoes the details are correct? Enter (y/n): ')
        if y_or_n.lower() == 'y':
            # Get the next available row
            next_row = ws.max_row + 1

            # Set values for the new row
            font = Font(name="Arial", size=10)
            alignment = Alignment(horizontal='center', vertical='center')

            cells = [
                (1, datetime_number),
                (2, company_name),
                (3, address_name),
                (4, TIN),
                (5, description_name),
                (6, TotalAmountPaid),
                (7, InputVat),
                (8, CPS),
            ]

            for column, value in cells:
                cell = ws.cell(row=next_row, column=column)
                cell.value = value
                cell.font = font
                cell.alignment = alignment
            # Save the workbook
            wb.save(Path)
            
            print("\n\033[92mData has been successfully saved.\033[00m")
            break
        elif y_or_n.lower() == 'n':
            Input_details(input("Enter the date (052424 = May,24,2024): "))
            Again()
            break
            
        else:
            
            print("\nInvalid key. Plase enter (y/n)\n")

    

if __name__ == "__main__":
    wb, Path = Requirements()
    Input_details(input("Enter the date (052424 = May,24,2024): "))
    Again()
